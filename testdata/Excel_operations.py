import openpyxl


class Excel_operation:
    # Give the location of the file
    path = "/Users/vasanp/PycharmProjects/pythonProject/selenium_mainassignment/testdata/application_data.xlsx"
    book = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute

    def get_row_count(self, sheet_name):
        sheet = self.book['users']
        return sheet.max_row

    def get_column_count(self, sheet_name):
        sheet = self.book[sheet_name]
        return sheet.max_column

    def read_data(self, sheet_name, row_num, column_num):
        sheet = self.book[sheet_name]
        return sheet.cell(row=row_num, column=column_num).value


    def write_data(self, sheet_name, row_num, column_num, data):
        sheet = self.book[sheet_name]
        sheet.cell(row=row_num, column=column_num).value = data
